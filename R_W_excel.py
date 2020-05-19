from openpyxl import  load_workbook
from openpyxl import  Workbook

#读取数据--先写散装代码--再写函数

def read_data(file_name,sheet_name):
    #打开一个Excel
    wb=load_workbook(file_name)

    #定位表单
    ws=wb[sheet_name]

    #读取所有的行数以及列数
    rows=ws.max_row
    cols=ws.max_column


    test_cases=[]
    #读取第一个测试用例的数据
    for i in range(2,rows+1):
        test_case=[]
        for j in range(1,cols-1):
          test_case.append(ws.cell(i,j).value)
        test_cases.append(test_case)
    return test_cases

if __name__ == '__main__':

    #获取注册的测试用例数据
    register_test_cases=read_data("test_case.xlsx",'register')
    print("注册的用例数据：{}".format(register_test_cases))

    #获取登陆的测试用例数据
    login_test_cases=read_data("test_case.xlsx",'login')
    print("注册的用例数据：{}".format(login_test_cases))

    #获取充值的测试用例数据
    recharge_test_cases=read_data("test_case.xlsx",'recharge')
    print("注册的用例数据：{}".format(recharge_test_cases))


#写入数据
def write_data(file_name,sheet_name,row,col,new_value):
    wb=load_workbook(file_name)
    ws=wb[sheet_name]

    ws.cell(row,col).value=new_value
    wb.save(file_name)